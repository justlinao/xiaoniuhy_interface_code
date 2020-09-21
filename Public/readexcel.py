# coding=utf-8
import openpyxl


class ReadExcel:

    def load_excel(self):
        """
        加载excel
        :return: wb
        """
        wb = openpyxl.load_workbook(r'C:\Users\linao\Documents\WXWork\1688851854143204\Cache\File\2020-09\飞鱼清理打包规则.xlsx')
        return wb

    def get_sheets_data(self, index=None):
        """
        加载sheet里的数据
        :return:
        """
        sheet_name = self.load_excel().sheetnames  # 读取工作簿的所有表单
        # print(sheet_name) 打印表单

        if index == None:  # 如果不传就默认拿第一个sheet
            index = 0
        data = self.load_excel()[sheet_name[index]]
        # print(data)
        return data

    def get_cell_value(self, row, cols):
        """
        获取单元格内容
        :return:
        """
        cell = self.get_sheets_data().cell(row=row, column=cols).value  # 根据具体的行列获取到对应单元格的值
        # print(cell)
        return cell

    def get_rows(self):
        """
        获取最大行
        :return:
        """
        maxrow = self.get_sheets_data().max_row
        # print(maxrow)
        return maxrow

    def get_row_value(self, row):
        """
        获取具体每一行的数据
        :param row: 传入你要的获取的行
        :return: 以列表储存行的数据
        """
        row_list = []
        data = self.get_sheets_data()[row]  # 这里只是拿到的对象
        for i in data:
            # print(i.value)
            row_list.append(i.value)
        return row_list

    def get_rows_value(self):
        """
        获取所有行的数据
        :return:
        """
        row_list = []
        rows = self.get_rows()  # 获取所有的行
        for i in range(rows-1):  # 这里的-1 是因为去除标题行因此要少遍历一行
            # print(i.value)
            row_list.append(self.get_row_value(i+2))  # 通过获取每一行的数据，获取所有行的数据
        return row_list

    def get_columns_value(self, key=None):
        """
        获取每一列的value
        :param key: 传入你想获取的列
        :return: 以列表储存列的数据
        """
        columns_list = []
        if key==None:
            key = 'A'
        columns_list_data = self.get_sheets_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        """
        获取行号
        :param case_id:
        :return:
        """
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num+1
        return num


if __name__ == '__main__':
    ReadExcel = ReadExcel()
    result = ReadExcel.get_columns_value('B')
    print(result)
    result.pop(0)
    print(result)
    with open(r'C:\Users\linao\Desktop\channel.txt', 'w+', encoding='utf-8') as f:
        for i in result:
            f.write('fyclean_hellogeek_'+i + '\n')

